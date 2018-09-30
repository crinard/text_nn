from __future__ import unicode_literals, print_function, division
import numpy
import pickle
import re
from sklearn.metrics import confusion_matrix
from sklearn.linear_model import LogisticRegression
from sklearn.linear_model import Perceptron
from sklearn.linear_model  import PassiveAggressiveClassifier
from sklearn.neural_network import MLPClassifier
from sklearn.svm import SVC
from sklearn.feature_extraction.text import CountVectorizer
import openpyxl
import numpy as np
import tqdm
import warnings
import random
import multiprocessing as mp
data = []
warnings.filterwarnings("ignore")
datafrom = openpyxl.load_workbook('NSCLCPD-1.xlsx')
train = datafrom.get_sheet_by_name("Training")
idarray = ['MSK_008', 'MSK_009', 'MSK_011', 'MSK_014', 'MSK_015', 'MSK_019', 'MSK_022', 'MSK_023', 'MSK_025', 'MSK_026', 'MSK_027', 'MSK_032', 'MSK_034', 'MSK_037', 'MSK_038', 'MSK_040', 'MSK_047', 'MSK_051', 'MSK_055', 'MSK_056', 'MSK_066', 'MSK_068', 'MSK_069', 'MSK_070', 'MSK_072', 'MSK_074', 'MSK_079', 'MSK_080', 'MSK_081', 'MSK_083', 'MSK_085', 'MSK_086', 'MSK_087', 'MSK_088', 'MSK_089', 'MSK_093', 'MSK_094', 'MSK_095', 'MSK_096', 'MSK_098', 'MSK_100', 'MSK_104', 'MSK_106', 'MSK_107', 'MSK_108', 'MSK_109', 'MSK_112', 'MSK_114', 'MSK_116', 'MSK_117', 'MSK_118', 'MSK_123', 'MSK_125', 'MSK_127', 'MSK_128', 'MSK_130', 'MSK_131', 'MSK_133', 'MSK_134', 'MSK_135', 'MSK_136', 'MSK_144', 'MSK_146', 'MSK_147', 'MSK_148', 'MSK_149', 'MSK_156', 'MSK_159', 'MSK_160', 'MSK_161', 'MSK_162', 'MSK_164', 'MSK_165', 'MSK_166', 'MSK_169', 'MSK_172', 'MSK_173', 'MSK_174', 'MSK_177', 'MSK_178', 'MSK_179', 'MSK_181', 'MSK_183', 'MSK_185', 'MSK_188', 'MSK_189', 'MSK_192', 'MSK_195', 'MSK_199', 'MSK_201', 'MSK_203', 'MSK_205', 'MSK_206', 'MSK_208', 'MSK_209', 'MSK_215', 'MSK_218', 'MSK_219', 'MSK_223', 'MSK_224', 'MSK_227', 'MSK_228', 'MSK_232', 'MSK_242', 'MSK_244', 'MSK_245', 'MSK_247', 'MSK_249', 'MSK_255', 'MSK_257', 'MSK_260', 'MSK_266', 'MSK_268', 'MSK_269', 'MSK_270', 'MSK_271', 'MSK_275', 'MSK_277', 'MSK_284', 'MSK_285', 'MSK_290', 'MSK_297', 'MSK_298', 'MSK_299', 'MSK_301', 'MSK_304', 'MSK_311', 'MSK_314', 'MSK_315', 'MSK_316', 'MSK_321', 'MSK_323', 'MSK_325', 'MSK_329', 'MSK_331', 'MSK_332', 'MSK_333', 'MSK_334', 'MSK_338', 'MSK_342', 'MSK_346', 'MSK_348', 'MSK_349', 'MSK_353', 'MSK_361', 'MSK_368', 'MSK_370', 'MSK_374', 'MSK_375', 'MSK_381', 'MSK_382', 'MSK_384', 'MSK_385', 'MSK_386', 'MSK_391', 'MSK_393', 'MSK_397', 'MSK_399', 'MSK_400', 'MSK_401', 'MSK_405', 'MSK_406', 'MSK_408', 'MSK_410', 'MSK_413', 'MSK_414', 'MSK_416', 'MSK_424', 'MSK_425', 'MSK_427', 'MSK_428', 'MSK_429', 'MSK_431', 'MSK_433', 'MSK_435', 'MSK_436', 'MSK_438', 'MSK_441', 'MSK_443', 'MSK_445', 'MSK_446', 'MSK_449', 'MSK_450', 'MSK_451', 'MSK_453', 'MSK_455', 'MSK_457', 'MSK_458', 'MSK_459', 'MSK_461', 'MSK_463', 'MSK_465', 'MSK_468', 'MSK_470', 'MSK_471', 'MSK_472', 'MSK_476', 'MSK_477', 'MSK_480', 'MSK_484', 'MSK_488', 'MSK_490', 'MSK_491', 'MSK_494', 'MSK_498', 'MSK_501', 'MSK_502', 'MSK_503', 'MSK_504', 'MSK_505', 'MSK_507', 'MSK_509', 'MSK_510', 'MSK_511', 'MSK_514', 'MSK_516', 'MSK_517', 'MSK_520', 'MSK_521', 'MSK_525', 'MSK_526', 'MSK_527', 'MSK_528', 'MSK_529', 'MSK_530', 'MSK_531', 'MSK_533', 'MSK_534', 'MSK_536', 'MSK_539', 'MSK_541', 'MSK_542', 'MSK_543', 'MSK_544', 'MSK_546', 'MSK_547', 'MSK_548', 'MSK_549', 'MSK_551', 'MSK_552', 'MSK_553', 'MSK_554', 'MSK_556', 'MSK_559', 'MSK_560', 'MSK_561', 'MSK_562', 'MSK_563', 'MSK_564', 'MSK_565', 'MSK_566', 'MSK_567', 'MSK_568', 'MSK_569', 'MSK_570', 'MSK_571', 'MSK_572', 'MSK_574', 'MSK_576', 'MSK_579', 'MSK_581', 'MSK_583', 'MSK_587', 'MSK_590', 'MSK_591', 'MSK_593', 'MSK_594', 'MSK_595', 'MSK_596', 'MSK_597', 'MSK_599', 'MSK_602', 'MSK_604', 'MSK_605', 'MSK_606', 'MSK_607', 'MSK_608', 'MSK_611', 'MSK_613', 'MSK_614', 'MSK_618', 'MSK_619', 'MSK_622', 'MSK_623', 'MSK_624', 'MSK_626', 'MSK_629', 'MSK_630', 'MSK_631', 'MSK_632', 'MSK_633', 'MSK_634', 'MSK_635', 'MSK_636', 'MSK_637', 'MSK_638', 'MSK_639', 'MSK_640', 'MSK_641', 'MSK_642', 'MSK_643', 'MSK_644', 'MSK_645', 'MSK_646', 'MSK_647', 'MSK_648', 'MSK_649', 'MSK_650', 'MSK_651', 'MSK_653', 'MSK_654', 'MSK_655', 'MSK_656', 'MSK_657', 'MSK_658', 'MSK_659', 'MSK_663', 'MSK_665', 'MSK_666', 'MSK_667', 'MSK_668', 'MSK_669', 'MSK_670', 'MSK_671', 'MSK_672', 'MSK_673', 'MSK_674', 'MSK_675', 'MSK_677', 'MSK_678', 'MSK_680', 'MSK_681', 'MSK_682', 'MSK_683', 'MSK_684', 'MSK_685', 'MSK_687', 'MSK_688', 'MSK_695', 'MSK_696', 'MSK_697', 'MSK_698', 'MSK_699', 'MSK_700', 'MSK_702', 'MSK_703', 'MSK_706', 'MSK_707', 'MSK_708', 'MSK_709', 'MSK_712', 'MSK_713', 'MSK_714', 'MSK_715', 'MSK_719', 'MSK_722', 'MSK_723', 'MSK_724', 'MSK_725', 'MSK_726', 'MSK_727', 'MSK_729']
# = [("","") for i in range(len(idarray))]
for indx in tqdm.tqdm(range(2,2252)):
        pat_id = train.cell(row = indx, column = 1).value
        for num in range(len(idarray)):
                idname = idarray[num]
                if (pat_id == idname):
                        text = train.cell(row = indx, column = 12).value
                        raw_label = train.cell(row = indx, column =3).value
                        if(raw_label in ("POD","SD","POD/brain")):
                                label = "0"
                        elif(raw_label in ("PR","CR")):
                                label = "1"
                        if(data[num][1] in ("0","1")):
                                before = data[num][0]
                                prior = data[num][1]
                                toadd = before+text
                                data[num] = (toadd,prior)
                        else:
                                data[num] = (text,label)
random.shuffle(data)
trainData = []
trainText = []
trainY = []
devData = []
devX = []
devY =[]
testText = []
testY = []
testData = []for num in range (8,700):
trainData = [data[j] for j in range(int(len(data)*.9))]
testData = [data[u] for u in range(int(len(data)*.9),(int(len(data))))]
trainText, trainY = [d[0] for d in trainData], [d[1] for d in trainData]

