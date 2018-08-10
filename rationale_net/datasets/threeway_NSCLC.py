import gzip
import re
import tqdm
from rationale_net.utils.embedding import get_indices_tensor
from rationale_net.datasets.factory import RegisterDataset
from rationale_net.datasets.abstract_dataset import AbstractDataset
from sklearn.datasets import fetch_20newsgroups
import openpyxl
import random
random.seed(0)


SMALL_TRAIN_SIZE = 800
datafrom = openpyxl.load_workbook('/scratch1/crinard/text_nn/rationale_net/datasets/NSCLCPD-1.xlsx')
train = datafrom.get_sheet_by_name("Training")
fuku = []

for indx in range(2,2252):
	a=train.cell(row = indx, column=12).value
	b=train.cell(row = indx, column=3).value
	if (b ==  "POD"):
		label1 = 0
		label_name1 = ('POD','POD/Brain')
		fuku.append((a,label1,label_name1))
	elif(b == 'SD') :
		label1 = 1
		label_name1 = ('SD')
		fuku.append((a,label1,label_name1))
	elif(b in ('PR','CR')):
		label1 = 2
		label_name1 = ('PR or CR')
		fuku.append((a,label1,label_name1))
	else:
		print(("loading error",b))
random.shuffle(fuku)
num_train = int(len(fuku)*.8)
num_dev = int(len(fuku)*.9)
traindata = fuku[:num_train]
devdata = fuku[num_train:num_dev]
testdata = fuku[num_dev:]
def preprocess_data(name):
	processed_data = []
	for i in range(len(name)):
		text = re.sub('\W+', ' ', name[i][0]).lower().strip()
		label = name[i][1]
		label_name= name[i][2]
		processed_data.append((text,label,label_name))
	return processed_data
@RegisterDataset('threeway_NSCLC')
class threeway_NSCLC(AbstractDataset):
	def __init__(self, args, word_to_indx, name, max_length=600):
		self.args = args
		self.args.num_class = 20
		self.name = name
		self.dataset = []
		self.word_to_indx  = word_to_indx
		self.max_length = max_length
		self.class_balance = {}
		if name in ['train']:
			print("train data")
			data = preprocess_data(traindata)
		elif name in ['dev']:
			print('dev data')
			data = preprocess_data(devdata)
		else:
			print('test data')
			data = preprocess_data(testdata)
		for indx, _sample in tqdm.tqdm(enumerate(data)):
			sample = self.processLine(_sample)
			if not sample['y'] in self.class_balance:
				self.class_balance[ sample['y'] ] = 0
			self.class_balance[ sample['y'] ] += 1
			self.dataset.append(sample)
			
		print ("Class balance", self.class_balance)

		if args.class_balance:
			raise NotImplementedError("NewsGroup dataset doesn't support balanced sampling")
		if args.objective == 'mse':
			raise NotImplementedError("News Group does not support Regression objective")

    ## Convert one line from beer dataset to {Text, Tensor, Labels}
	def processLine(self, row):
		text, label, label_name = row[0] , row[1], row[2]
		text = " ".join(text.split()[:self.max_length])
		x =  get_indices_tensor(text.split(), self.word_to_indx, self.max_length)
		sample = {'text':text,'x':x, 'y':label, 'y_name': label_name}
		return sample
